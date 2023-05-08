import time
import cv2
import numpy as np
import os
from gtts import gTTS
import pyttsx3
import keyboard
import openpyxl
from datetime import datetime


recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read('trainer/trainer.yml')
cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath);
font = cv2.FONT_HERSHEY_SIMPLEX
id = 0
# index 1 : umang, Index 3: Ujjwal
names = ['Umang', 'Maneesh Chauhan','Vansh Choudhary']
truefalse = [True,True,True]
# Initialize and start realtime video capture
cam = cv2.VideoCapture(0)
cam.set(3, 640)
cam.set(4, 480)
minW = 0.1 * cam.get(3)
minH = 0.1 * cam.get(4)
engine = pyttsx3.init()
n = []
p = 0
temp = ""
while True:
    ret, frame = cam.read()
    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2HSV)
    contrast = 1.25
    brightness = 50
    frame[:, :, 2] = np.clip(contrast * frame[:, :, 2] + brightness, 0, 255)
    frame = cv2.cvtColor(frame, cv2.COLOR_HSV2BGR)
    ret, img = cam.read()
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)

    faces = faceCascade.detectMultiScale(
        gray,
        scaleFactor=2,
        minNeighbors=5,
        minSize=(int(minW), int(minH)),
    )

    for (x, y, w, h) in faces:
        #cv2.rectangle(img, (x, y), (x + w, y + h), (0, 255, 0), 2)
        id, confidence = recognizer.predict(gray[y:y + h, x:x + w])


        if (confidence < 50):
            if (truefalse[id]):
                engine.say(names[id] + ' is present today')
                engine.runAndWait()
                n.append(names[id])
                truefalse[id] = False
                p += 1
            temp = names[id]
            confidence = "  {0}%".format(round(100 - confidence))
        elif (confidence<60):
            temp = names[id]
            confidence = "  {0}%".format(round(100 - confidence))
        else:
            temp = "unknown"
            confidence = "  {0}%".format(round(100 - confidence))

        cv2.putText(
            img,
            str(temp),
            (x+40, y),
            font,
            1,
            (0, 255, 0),
            2
        )
        # cv2.putText(
        #      img,
        #      str(confidence),
        #      (x + 5, y + h - 5),
        #      font,
        #      1,
        #      (0, 0, 0),
        #      1
        #  )
    a = 0

    #for printing the data in console
    if keyboard.is_pressed("p"):
        print("Student's present : ")
        for x in n:
            print(x)

        print('\n')
        print("Student's Absent : ")
        for x in names:
            if x  not in n:
                a += 1
                print(x)

        print('\n')
        print('Total students : ' + str(len(names)))
        print('Present : ' + str(p))
        print('Absent : ' + str(a))


        #for updating the excel file 
        #path = r'D:\Face-Detection-with-Name-Recognition-main\attendance.xlsx'
        workbook = openpyxl.load_workbook("attendance.xlsx")
        sheet = workbook.active

        now = datetime.now()
        date = now.strftime("%d/%m/%Y")
        c1 = sheet.cell(row = 1,column = 3)
        column = 3
        for i in range(1,10):
            col = sheet.cell(row = 1,column = i+2).value
            if(col == date):
                break
            elif(col == None):
                c1 = sheet.cell(row = 1,column = i+2)
                column = i+2
                c1.value = date
                break

        index = 0
        for x in truefalse:
            c2 = sheet.cell(row = 2 + index, column = column)
            if(x == False):
                c2.value = 'P'
            else:
                c2.value = 'A'
            index += 1

        workbook.save("attendance.xlsx")
        print("Updated the attendance of " + date)
        
        break
    cv2.imshow('camera', img)
    k = cv2.waitKey(10) & 0xff
    if k == 27:
        break

        
cam.release()
cv2.destroyAllWindows()